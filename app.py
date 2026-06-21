from flask import Flask, send_from_directory, send_file, request, redirect
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import requests

app = Flask(__name__)

BASE_URL = "https://amelia90.onrender.com"
MAPS_LINK = "https://maps.app.goo.gl/T3QYpdr4MEtbXsJ96"
ADMIN_SECRET = "8F4K2X9P7Q"

SUPABASE_URL = "https://oqfokoxdnwvfbkvzeqew.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9xZm9rb3hkbnd2ZmJrdnplcWV3Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODA2MDc5NDEsImV4cCI6MjA5NjE4Mzk0MX0.JuC0RPkYZlUJM7itCVOAQCV_-XFv2v2ZkOVqeMe3Q-c"


def supabase_headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }


def supabase_get(tabela, query=""):
    url = f"{SUPABASE_URL}/rest/v1/{tabela}{query}"
    r = requests.get(url, headers=supabase_headers())
    r.raise_for_status()
    return r.json()


def supabase_post(tabela, dados):
    url = f"{SUPABASE_URL}/rest/v1/{tabela}"
    r = requests.post(url, headers=supabase_headers(), json=dados)
    r.raise_for_status()
    return r.json()


def supabase_patch(tabela, filtro, dados):
    url = f"{SUPABASE_URL}/rest/v1/{tabela}{filtro}"
    r = requests.patch(url, headers=supabase_headers(), json=dados)
    r.raise_for_status()
    return r.json()


def supabase_delete(tabela, filtro):
    url = f"{SUPABASE_URL}/rest/v1/{tabela}{filtro}"
    r = requests.delete(url, headers=supabase_headers())
    r.raise_for_status()


@app.route("/")
def home():
    return """
    <h1>90 Anos da Amélia</h1>
    <p>Sistema RSVP funcionando!</p>
    """


@app.route("/uploads/<nome_arquivo>")
def uploads(nome_arquivo):
    return send_from_directory("uploads", nome_arquivo)


@app.route("/admin/convidados")
def convidados_publico():
    return "<h1>Acesso restrito.</h1>"


@app.route("/admin/<segredo>/convidados")
def listar_convidados(segredo):
    if segredo != ADMIN_SECRET:
        return "<h1>Acesso negado.</h1>"

    convidados = supabase_get("convidados", "?select=*&order=nome.asc")

    html = f"""
    <h1>Links dos Convidados</h1>

    <p><a href="/admin/{segredo}">Voltar ao painel</a></p>

    <table border="1" cellpadding="8">
        <tr>
            <th>ID</th>
            <th>Código</th>
            <th>Nome</th>
            <th>Telefone</th>
            <th>Limite</th>
            <th>Link RSVP</th>
        </tr>
    """

    for c in convidados:
        link = f"{BASE_URL}/rsvp/{c['codigo']}"
        html += f"""
        <tr>
            <td>{c['id']}</td>
            <td>{c['codigo']}</td>
            <td>{c.get('nome', '')}</td>
            <td>{c.get('telefone', '')}</td>
            <td>{c.get('limite', 1)}</td>
            <td><a href="{link}" target="_blank">{link}</a></td>
        </tr>
        """

    html += "</table>"
    return html


@app.route("/rsvp/<codigo>")
def rsvp(codigo):
    convidados = supabase_get("convidados", f"?codigo=eq.{codigo}&select=*")

    if not convidados:
        return "<h1>Convite não encontrado.</h1>"

    convidado = convidados[0]
    nome = convidado.get("nome", "")

    return f"""
    <!DOCTYPE html>
    <html lang="pt-br">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>90 Anos da Amélia</title>
        <style>
            body {{
                margin: 0;
                font-family: Arial, sans-serif;
                background: #fff4ec;
                color: #3b1f14;
                text-align: center;
            }}
            .container {{
                max-width: 720px;
                margin: 0 auto;
                padding: 20px;
            }}
            .convite {{
                width: 100%;
                max-width: 600px;
                border-radius: 16px;
                box-shadow: 0 4px 18px rgba(0,0,0,0.18);
                margin-bottom: 24px;
            }}
            .card {{
                background: #ffffff;
                border-radius: 18px;
                padding: 24px;
                box-shadow: 0 4px 16px rgba(0,0,0,0.12);
                margin-bottom: 24px;
            }}
            h1 {{
                color: #7a3b22;
            }}
            p {{
                font-size: 18px;
                line-height: 1.5;
            }}
            .botao {{
                display: inline-block;
                background: #c46b4f;
                color: white;
                text-decoration: none;
                padding: 14px 22px;
                border-radius: 12px;
                font-weight: bold;
                margin: 8px;
            }}
            .botao-secundario {{
                background: #8a5a44;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <img src="/uploads/convite.png" class="convite" alt="Convite 90 anos da Amélia">

            <div class="card">
                <h1>Olá, {nome}!</h1>

                <p>
                    Ficamos muito felizes em convidar você para a comemoração dos
                    <strong>90 anos da Dona Amélia</strong>.
                </p>

                <p>
                    Para confirmar sua presença, clique no botão abaixo e informe
                    os nomes completos das pessoas que participarão.
                </p>

                <a class="botao" href="/rsvp/{codigo}/confirmar">
                    Confirmar Presença
                </a>

                <a class="botao botao-secundario" href="{MAPS_LINK}" target="_blank">
                    📍 Como Chegar
                </a>
            </div>
        </div>
    </body>
    </html>
    """


@app.route("/rsvp/<codigo>/confirmar", methods=["GET", "POST"])
def confirmar(codigo):
    convidados = supabase_get("convidados", f"?codigo=eq.{codigo}&select=*")

    if not convidados:
        return "<h1>Convite não encontrado.</h1>"

    convidado = convidados[0]
    convidado_id = convidado["id"]
    nome_convidado = convidado.get("nome", "")
    respondeu = convidado.get("respondeu", False)
    limite = int(convidado.get("limite") or 1)

    if respondeu:
        return """
        <h1>Confirmação já registrada</h1>
        <p>Sua solicitação já foi registrada anteriormente.</p>
        <p>
        Caso necessite de alguma alteração, entre em contato diretamente
        com a organização do evento.
        </p>
        """

    if request.method == "POST":
        observacoes = request.form.get("observacoes", "")

        nomes = request.form.getlist("nome[]")
        tipos = request.form.getlist("tipo[]")
        idades = request.form.getlist("idade[]")

        nomes_validos = [n.strip() for n in nomes if n.strip()]

        if len(nomes_validos) > limite:
            return f"""
            <h1>Limite excedido</h1>
            <p>Este convite permite informar até {limite} participante(s).</p>
            <p>Volte e ajuste a lista antes de enviar.</p>
            """

        for nome, tipo, idade in zip(nomes, tipos, idades):
            if tipo == "Crianca" and not str(idade).strip():
                return """
                <h1>Idade obrigatória</h1>
                <p>Para crianças, informe a idade antes de enviar.</p>
                <p>Volte e ajuste o formulário.</p>
                """

        confirmacao = supabase_post("confirmacoes", {
            "convidado_id": convidado_id,
            "observacoes": observacoes
        })[0]

        confirmacao_id = confirmacao["id"]

        participantes = []

        for nome, tipo, idade in zip(nomes, tipos, idades):
            nome = nome.strip()

            if not nome:
                continue

            idade_final = None

            if tipo == "Crianca" and idade:
                idade_final = int(idade)

            participantes.append({
                "confirmacao_id": confirmacao_id,
                "nome": nome,
                "tipo": tipo,
                "idade": idade_final
            })

        if participantes:
            supabase_post("participantes", participantes)

        supabase_patch(
            "convidados",
            f"?id=eq.{convidado_id}",
            {"respondeu": True}
        )

        return """
        <h1>Solicitação recebida com sucesso!</h1>
        <p>
        Agradecemos seu interesse em participar da comemoração dos
        90 anos da Dona Amélia.
        </p>
        <p>
        Os nomes informados foram encaminhados ao cerimonial e serão analisados
        conforme a organização do evento.
        </p>
        <p>
        Após a validação, entraremos em contato para informar os nomes
        efetivamente confirmados.
        </p>
        """

    return f"""
    <!DOCTYPE html>
    <html lang="pt-br">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Confirmar Presença</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                background: #fff4ec;
                color: #3b1f14;
                max-width: 900px;
                margin: auto;
                padding: 20px;
            }}
            .card {{
                background: white;
                padding: 24px;
                border-radius: 18px;
                box-shadow: 0 4px 16px rgba(0,0,0,0.12);
            }}
            .participante {{
                border: 1px solid #ddd;
                padding: 15px;
                margin-bottom: 12px;
                border-radius: 10px;
                background: #fffaf7;
            }}
            input, select, textarea {{
                width: 100%;
                padding: 10px;
                margin-top: 5px;
                margin-bottom: 12px;
                box-sizing: border-box;
            }}
            button {{
                background: #c46b4f;
                color: white;
                border: none;
                padding: 12px 18px;
                border-radius: 10px;
                cursor: pointer;
                font-weight: bold;
            }}
            .secundario {{
                background: #8a5a44;
            }}
        </style>

        <script>
            const LIMITE = {limite};

            function adicionarParticipante() {{
                let container = document.getElementById("participantes");
                let total = container.getElementsByClassName("participante").length;

                if (total >= LIMITE) {{
                    alert("Este convite permite informar até " + LIMITE + " participante(s).");
                    return;
                }}

                let bloco = document.createElement("div");
                bloco.className = "participante";

                bloco.innerHTML = `
                    <label>Nome Completo</label>
                    <input type="text" name="nome[]" required>

                    <label>Tipo</label>
                    <select name="tipo[]">
                        <option value="Adulto">Adulto</option>
                        <option value="Crianca">Criança</option>
                    </select>

                    <label>Idade (somente criança)</label>
                    <input type="number" name="idade[]" min="0">
                `;

                container.appendChild(bloco);
            }}
        </script>
    </head>

    <body>
        <div class="card">
            <h1>Participantes da Família</h1>

            <p>
                Convite vinculado a:
                <strong>{nome_convidado}</strong>
            </p>

            <p>
                Este convite permite informar até
                <strong>{limite}</strong> participante(s).
            </p>

            <form method="POST">
                <div id="participantes">
                    <div class="participante">
                        <label>Nome Completo</label>
                        <input type="text" name="nome[]" required>

                        <label>Tipo</label>
                        <select name="tipo[]">
                            <option value="Adulto">Adulto</option>
                            <option value="Crianca">Criança</option>
                        </select>

                        <label>Idade (somente criança)</label>
                        <input type="number" name="idade[]" min="0">
                    </div>
                </div>

                <button type="button" class="secundario" onclick="adicionarParticipante()">
                    ➕ Adicionar Participante
                </button>

                <br><br>

                <label>Observações</label>
                <textarea name="observacoes" rows="5"></textarea>

                <br><br>

                <button type="submit">
                    ENVIAR SOLICITAÇÃO
                </button>
            </form>
        </div>
    </body>
    </html>
    """


@app.route("/admin/<segredo>")
def admin(segredo):
    if segredo != ADMIN_SECRET:
        return "<h1>Acesso negado.</h1>"

    convidados = supabase_get("convidados", "?select=*&order=nome.asc")
    confirmacoes = supabase_get("confirmacoes", "?select=*")
    participantes = supabase_get("participantes", "?select=*")

    total_convidados = len(convidados)
    total_responderam = len([c for c in convidados if c.get("respondeu")])
    total_pendentes = total_convidados - total_responderam
    total_participantes = len(participantes)
    total_adultos = len([p for p in participantes if p.get("tipo") == "Adulto"])
    total_criancas = len([p for p in participantes if p.get("tipo") == "Crianca"])

    confirmacoes_por_convidado = {}
    for cf in confirmacoes:
        confirmacoes_por_convidado.setdefault(cf["convidado_id"], []).append(cf["id"])

    participantes_por_confirmacao = {}
    for p in participantes:
        participantes_por_confirmacao.setdefault(p["confirmacao_id"], []).append(p)

    html = f"""
    <h1>Painel Administrativo - 90 Anos da Amélia</h1>

    <h2>Resumo</h2>
    <ul>
        <li>Total de convites: {total_convidados}</li>
        <li>Responderam: {total_responderam}</li>
        <li>Pendentes: {total_pendentes}</li>
        <li>Participantes informados: {total_participantes}</li>
        <li>Adultos informados: {total_adultos}</li>
        <li>Crianças informadas: {total_criancas}</li>
    </ul>

    <p>
        <a href="/admin/{segredo}/convidados">Ver links dos convidados</a>
    </p>

    <p>
        <a href="/admin/{segredo}/exportar">
            Exportar Lista para Excel
        </a>
    </p>

    <table border="1" cellpadding="8">
        <tr>
            <th>Convidado</th>
            <th>Telefone</th>
            <th>Limite</th>
            <th>Confirmados / Limite</th>
            <th>Vagas restantes</th>
            <th>Status</th>
            <th>Ações</th>
        </tr>
    """

    for c in convidados:
        convidado_id = c["id"]
        limite = int(c.get("limite") or 1)

        confirmacao_ids = confirmacoes_por_convidado.get(convidado_id, [])
        total = 0
        for cf_id in confirmacao_ids:
            total += len(participantes_por_confirmacao.get(cf_id, []))

        restantes = max(limite - total, 0)

        if not c.get("respondeu"):
            status = "Pendente"
        elif total >= limite:
            status = "Completo"
        else:
            status = "Parcial"

        html += f"""
        <tr>
            <td>{c.get('nome', '')}</td>
            <td>{c.get('telefone', '')}</td>
            <td>{limite}</td>
            <td>{total} / {limite}</td>
            <td>{restantes}</td>
            <td>{status}</td>
            <td>
                <a href="/admin/{segredo}/convidado/{convidado_id}">
                    Ver detalhes
                </a>
                |
                <a href="/admin/{segredo}/convidado/{convidado_id}/limite">
                    Alterar limite
                </a>
                |
                <a href="/admin/{segredo}/convidado/{convidado_id}/resetar"
                   onclick="return confirm('Tem certeza que deseja limpar a confirmação deste convidado?');">
                    Limpar confirmação
                </a>
            </td>
        </tr>
        """

    html += "</table>"
    return html


@app.route("/admin/<segredo>/convidado/<int:convidado_id>")
def detalhes_convidado(segredo, convidado_id):
    if segredo != ADMIN_SECRET:
        return "<h1>Acesso negado.</h1>"

    convidados = supabase_get("convidados", f"?id=eq.{convidado_id}&select=*")

    if not convidados:
        return "<h1>Convidado não encontrado.</h1>"

    convidado = convidados[0]

    confirmacoes = supabase_get(
        "confirmacoes",
        f"?convidado_id=eq.{convidado_id}&select=*&order=id.desc"
    )

    if not confirmacoes:
        return f"""
        <h1>Este convidado ainda não respondeu.</h1>

        <p><strong>Convidado:</strong> {convidado.get('nome', '')}</p>
        <p><strong>Telefone:</strong> {convidado.get('telefone', '')}</p>
        <p><strong>Limite:</strong> {convidado.get('limite', 1)}</p>

        <p>
            <a href="/admin/{segredo}/convidado/{convidado_id}/limite">
                Alterar limite
            </a>
        </p>

        <p>
            <a href="/admin/{segredo}">Voltar ao painel</a>
        </p>
        """

    confirmacao = confirmacoes[0]
    confirmacao_id = confirmacao["id"]

    participantes = supabase_get(
        "participantes",
        f"?confirmacao_id=eq.{confirmacao_id}&select=*&order=id.asc"
    )

    total_participantes = len(participantes)
    limite = int(convidado.get("limite") or 1)
    restantes = max(limite - total_participantes, 0)

    html = f"""
    <h1>Detalhes da Confirmação</h1>

    <p><strong>Convidado:</strong> {convidado.get('nome', '')}</p>
    <p><strong>Telefone:</strong> {convidado.get('telefone', '')}</p>
    <p><strong>Limite:</strong> {limite}</p>
    <p><strong>Confirmados:</strong> {total_participantes} / {limite}</p>
    <p><strong>Vagas restantes:</strong> {restantes}</p>
    <p><strong>Data:</strong> {confirmacao.get('data_confirmacao', '')}</p>

    <h2>Participantes informados</h2>
    <ul>
    """

    for p in participantes:
        if p.get("tipo") == "Crianca" and p.get("idade"):
            html += f"<li>{p.get('nome', '')} - Criança, {p.get('idade')} anos</li>"
        else:
            html += f"<li>{p.get('nome', '')} - Adulto</li>"

    html += f"""
    </ul>

    <h2>Observações</h2>
    <p>{confirmacao.get('observacoes') or "Nenhuma observação."}</p>

    <p>
        <a href="/admin/{segredo}/convidado/{convidado_id}/limite">
            Alterar limite
        </a>
    </p>

    <p>
        <a href="/admin/{segredo}/convidado/{convidado_id}/resetar"
           onclick="return confirm('Tem certeza que deseja limpar esta confirmação?');">
            Limpar confirmação deste convidado
        </a>
    </p>

    <p>
        <a href="/admin/{segredo}">Voltar ao painel</a>
    </p>
    """

    return html


@app.route("/admin/<segredo>/convidado/<int:convidado_id>/limite", methods=["GET", "POST"])
def alterar_limite(segredo, convidado_id):
    if segredo != ADMIN_SECRET:
        return "<h1>Acesso negado.</h1>"

    convidados = supabase_get("convidados", f"?id=eq.{convidado_id}&select=*")

    if not convidados:
        return "<h1>Convidado não encontrado.</h1>"

    convidado = convidados[0]

    if request.method == "POST":
        novo_limite = request.form.get("limite", "1")

        try:
            novo_limite = int(novo_limite)
        except ValueError:
            novo_limite = 1

        if novo_limite < 1:
            novo_limite = 1

        supabase_patch(
            "convidados",
            f"?id=eq.{convidado_id}",
            {"limite": novo_limite}
        )

        return redirect(f"/admin/{segredo}")

    return f"""
    <h1>Alterar Limite</h1>

    <p><strong>Convidado:</strong> {convidado.get('nome', '')}</p>
    <p><strong>Telefone:</strong> {convidado.get('telefone', '')}</p>
    <p><strong>Limite atual:</strong> {convidado.get('limite', 1)}</p>

    <form method="POST">
        <label>Novo limite de participantes:</label><br>
        <input type="number" name="limite" min="1" value="{convidado.get('limite', 1)}" required>
        <br><br>
        <button type="submit">Salvar novo limite</button>
    </form>

    <p>
        <a href="/admin/{segredo}">Voltar ao painel</a>
    </p>
    """


@app.route("/admin/<segredo>/convidado/<int:convidado_id>/resetar")
def resetar_convidado(segredo, convidado_id):
    if segredo != ADMIN_SECRET:
        return "<h1>Acesso negado.</h1>"

    confirmacoes = supabase_get(
        "confirmacoes",
        f"?convidado_id=eq.{convidado_id}&select=*"
    )

    for cf in confirmacoes:
        supabase_delete("participantes", f"?confirmacao_id=eq.{cf['id']}")

    supabase_delete("confirmacoes", f"?convidado_id=eq.{convidado_id}")

    supabase_patch(
        "convidados",
        f"?id=eq.{convidado_id}",
        {"respondeu": False}
    )

    return redirect(f"/admin/{segredo}")


@app.route("/admin/<segredo>/exportar")
def exportar_excel(segredo):
    if segredo != ADMIN_SECRET:
        return "<h1>Acesso negado.</h1>"

    convidados = supabase_get("convidados", "?select=*")
    confirmacoes = supabase_get("confirmacoes", "?select=*")
    participantes = supabase_get("participantes", "?select=*")

    convidados_por_id = {c["id"]: c for c in convidados}
    confirmacoes_por_id = {cf["id"]: cf for cf in confirmacoes}

    lista_completa = []
    criancas = []

    for p in participantes:
        cf = confirmacoes_por_id.get(p["confirmacao_id"], {})
        convidado = convidados_por_id.get(cf.get("convidado_id"), {})

        linha = {
            "Nome": p.get("nome", ""),
            "Tipo": p.get("tipo", ""),
            "Idade": p.get("idade") or "",
            "Convidado de origem": convidado.get("nome", ""),
            "Data da confirmação": cf.get("data_confirmacao", "")
        }

        lista_completa.append(linha)

        if p.get("tipo") == "Crianca":
            criancas.append({
                "Nome": p.get("nome", ""),
                "Idade": p.get("idade") or "",
                "Convidado de origem": convidado.get("nome", "")
            })

    caminho = "exportacoes/lista_portaria.xlsx"

    colunas_lista = [
        "Nome",
        "Tipo",
        "Idade",
        "Convidado de origem",
        "Data da confirmação"
    ]

    colunas_criancas = [
        "Nome",
        "Idade",
        "Convidado de origem"
    ]

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        pd.DataFrame(lista_completa, columns=colunas_lista).to_excel(
            writer,
            sheet_name="Lista Portaria",
            index=False
        )

        pd.DataFrame(criancas, columns=colunas_criancas).to_excel(
            writer,
            sheet_name="Crianças",
            index=False
        )

    workbook = load_workbook(caminho)

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 3

    workbook.save(caminho)

    return send_file(
        caminho,
        as_attachment=True,
        download_name="lista_portaria_amelia90.xlsx"
    )


if __name__ == "__main__":
    app.run(debug=True)